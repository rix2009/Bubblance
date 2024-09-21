from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login, authenticate, logout
from django.contrib import messages
from collections import defaultdict
from django.utils import timezone
import time
import win32com.client
import pythoncom
import os
from concurrent.futures import ThreadPoolExecutor, as_completed
from django.db.models import Q
from reportlab.lib import colors, pagesizes
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
import pytz
import logging
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from django.conf import settings
from django.http import JsonResponse, HttpResponseRedirect, HttpResponse
from django.views.generic.edit import FormView, CreateView
from django.views.generic.base import TemplateView
from django.utils.decorators import method_decorator
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.models import User
from django.urls import reverse, path
from datetime import datetime, timedelta, date
from .forms import UpdateInstitutionForm, NewUserForm, NewAmbulanceForm, EqupmentInAmbulanceForm, NewCrewForm
from .forms import UpdateUserForm, NewCustomerForm, NewInstitution, CustomerRequestForm, DateFilterForm
from .models import BUser, Ambulance, EqInAmbulance, AmbulanceCrew, Institution, Customer, CustomerRide, CustomerRequest
from Bubblance.mixins import AjaxFormMixin, FormErrors, RedrectParams
import requests

# Create your views here.

def index(request):
	context = {}
	context["users"] = BUser.objects.filter(status = 1)
	return render(request, "index.html", context)


def register_request(request):
	form = NewUserForm()
	if request.method == "POST":
		form = NewUserForm(request.POST)
		if form.is_valid():
			user = form.save()
			messages.success(request, "user added successfuly." )
			return redirect("drivers")
		messages.error(request, "Unsuccessful registration. Invalid information.")
	return render (request=request, template_name="register.html", context={"register_form":form})


def login_request(request):
    if request.method == "POST":
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                buser = BUser.objects.get(username=user)
                if buser.status == 1:
                    login(request, user)
                    messages.info(request, f"You are now logged in as {buser.firstname} {buser.lastname}.")
                    return redirect(reverse("home"), kwargs={"user": user})
                else:
                    messages.error(request,"The user is not active.")
            else:
                messages.error(request,"Invalid username or password.")
    else:
        messages.error(request,"Invalid username or password.")
        form = AuthenticationForm()
    return render(request=request, template_name="login.html", context={"login_form":form})


def logout_request(request):
	logout(request)
	messages.info(request, "You have successfully logged out.") 
	return redirect("home")


def create_ambulance(request):
	form = NewAmbulanceForm()
	if request.method == "POST":
		form = NewAmbulanceForm(request.POST)
		if form.is_valid():
			ambulance = form.save()
			messages.success(request, "Ambulance added successfuly." )
			return redirect("ambulance")
		messages.error(request, "Unsuccessful registration. Invalid information.")
	return render (request=request, template_name="create_ambulance.html", context={"create_ambulance_form":form})


def ambulance(request):
	context = {}
	context["no_amb"] = False
	context["ambulances"] = Ambulance.objects.filter(status = 1)
	if context["ambulances"].count() == 0:
		context["no_amb"] = True
	if request.method == "POST":
		amb = request.POST
		render(request=request, template_name="ambulance_info.html", context = {"amb":amb})
	return render (request=request, template_name="ambulance.html", context = context)


def create_equipment(request):
	amb = Ambulance.objects.get(ambulance_id = request.POST.get('amb'))
	to_submit = request.POST["form_status"]
	form = EqupmentInAmbulanceForm()
	if request.method == "POST" and to_submit:
		form = EqupmentInAmbulanceForm(request.POST)
		if form.is_valid():
			eq = form.save()
			messages.success(request, "Equipment added successfuly." )
			return ambulance_info(request)
		messages.error(request, "Unsuccessful registration. Invalid information.")
	return render (request=request, template_name="create_eq.html", 
				context={"form":form, "amb":amb,})


def drivers(request):
    show_inactive = request.GET.get('show_inactive', 'false') == 'true'
    
    if show_inactive:
        users = BUser.objects.filter(status=0)
    else:
        users = BUser.objects.filter(status=1)
    
    context = {
        "users": users,
        "no_user": users.count() == 0,
        "show_inactive": show_inactive
    }
    return render(request=request, template_name="drivers.html", context=context)


def ambulance_info(request):
    context = {}
    if request.method == "POST":
        amb = Ambulance.objects.get(ambulance_id = request.POST.get('amb'))
        context["amb"] = amb
        eq_in_amb = []
        for e in EqInAmbulance.objects.filter(am_id = amb):
            eq_in_amb.append(e)
        context["equipment"] = eq_in_amb
        no_driver = True
        for d in AmbulanceCrew.objects.filter(ambulance_id = amb.ambulance_id, start_time__lte = datetime.now()):
            if d.end_time == None :
                context["driver"] = BUser.objects.get(username = d.driver_id.username)
                no_driver = False
        if no_driver:
            context["busers"] = BUser.objects.all()
        context["no_driver"] = no_driver
        context["new_crew_form"] = NewCrewForm
        return render (request, "ambulance_info.html",context)
    return redirect (reverse("ambulance"))


def disable_ambulance(request):
	if request.method == "POST":
		amb = request.POST.get("amb")
		instance = Ambulance.objects.filter(ambulance_id = amb.ambulance_id)
		for crew in AmbulanceCrew.objects.filter(ambulance_id = amb.ambulance_id):
			crew.status = 2
		for eq in EqInAmbulance.objects.filter(am_id = amb.ambulance_id):
			eq.status = 2
		instance.status = 2
		return redirect (reverse("ambulance"))
	return redirect (reverse("ambulance"))


def driver_to_amb(request):
	if request.method == "POST":
		form = NewCrewForm(request.POST)
		if form.is_valid():
			crew = form.save()
			messages.success(request, "Driver added succesfully to the ambulance")
			return render(request=request, template_name="ambulance_info.html", context =
				  {"amb":form.cleaned_data.get("ambulance_id"), "driver":form.cleaned_data.get("driver_id")})
		messages.error(request, "Failed, pls try again with valid info")
	return redirect(reverse("ambulance"))


def end_crew_time(request):
	if request.method == "POST":
		amb = Ambulance.objects.get(ambulance_id=request.POST.get("ambulance_id"))
		user = BUser.objects.get(username=request.POST.get("username"))
		crew = AmbulanceCrew.objects.get(ambulance_id=amb, driver_id=user, end_time=None)
		crew.end_time = datetime.now()
		crew.save()
		return render(request=request, template_name="ambulance_info.html", context = {"amb":amb, "no_driver":True, "new_crew_form":NewCrewForm})
	return redirect(reverse("ambulance"))	


def driver_info(request):
    user = request.POST.get("buser")
    buser = get_object_or_404(BUser, username=user)

    if request.method == 'POST' and request.POST.get("save_form") == "True":
        form = UpdateUserForm(request.POST, instance=buser)
        if form.is_valid():
            form.save()
            messages.success(request, "User updated successfully.")
            return redirect("drivers")
        messages.error(request, "Unsuccessful update. Invalid information.")
    else:
        form = UpdateUserForm(instance=buser)
    
    return render(request, "driver_info.html", {"buser": buser, "update_user_form": form, "save_form": True})


def disable_driver(request):
	if request.method == "POST":
		user = request.POST.get("buser")
		instance = BUser.objects.filter(username = user.user.username)
		for crew in AmbulanceCrew.objects.filter(driver_id = user.user.username):
			crew.status = 2
		instance.status = 2
		return redirect (reverse("drivers"))
	return redirect (reverse("drivers"))


def edit_equipment(request):
	if request.method =="POST":
		amb = Ambulance.objects.get(ambulance_id=request.POST.get("amb"))
		for e in EqInAmbulance.objects.filter(am_id = amb.ambulance_id):
			amount = request.POST[str(e.eq_id)]
			if amount == '0':
				e.delete()
			else:
				e.amount = amount
				e.save()
	return ambulance_info(request)


def institutions(request):
    context = {}
    show_inactive = request.GET.get('show_inactive', 'false') == 'true'
    
    if show_inactive:
        context["inst"] = Institution.objects.filter(status=2, in_institution__isnull=True)
    else:
        context["inst"] = Institution.objects.filter(status=1, in_institution__isnull=True)
    
    context["no_inst"] = context["inst"].count() == 0
    context["show_inactive"] = show_inactive
    return render(request=request, template_name="institutions.html", context=context)


def customers(request):
	context = {}
	context["customers"] = Customer.objects.all()
	context["no_customer"] = False
	if context["customers"].count() == 0:
		context["no_customer"] = True
	return render (request=request, template_name="customers.html", context = context)


def add_institution(request):
    parent_id = request.GET.get('parent_id')
    initial_data = {}
    
    if parent_id:
        parent_institution = Institution.objects.get(inst_id=parent_id)
        initial_data['in_institution'] = parent_institution
        initial_data['institution_adress'] = parent_institution.institution_adress

    if request.method == "POST":
        form = NewInstitution(request.POST)
        if form.is_valid():
            institution = form.save(commit=False)
            if parent_id:
                institution.in_institution = parent_institution
                if not institution.institution_adress:
                    institution.institution_adress = parent_institution.institution_adress
            institution.save()
            messages.success(request, "Institution added successfully.")
            return redirect("institutions")
        messages.error(request, "Unsuccessful registration. Invalid information.")
    else:
        form = NewInstitution(initial=initial_data)
    
    return render(request=request, template_name="add_institution.html", context={"add_institution_form": form})


def institution_info(request):
    inst_id = request.POST.get("inst_id") or request.GET.get("inst_id")
    inst = get_object_or_404(Institution, inst_id=inst_id)
    child_institutions = Institution.objects.filter(in_institution=inst)
    
    if request.method == 'POST' and request.POST.get("save_form") == "True":
        form = UpdateInstitutionForm(request.POST, instance=inst)
        if form.is_valid():
            form.save()
            messages.success(request, "Institution updated successfully.")
            return redirect("institutions")
        else:
            messages.error(request, "Unsuccessful update. Invalid information.")
    else:
        form = UpdateInstitutionForm(instance=inst)
    
    return render(request, "institution_info.html", {
        "inst": inst,
        "update_inst_form": form,
        "save_form": True,
        "child_institutions": child_institutions
    })


def get_travel_time(api_key, origin, destination, departure_time):
    israel_tz = pytz.timezone('Asia/Jerusalem')
    reference_time = datetime.now(israel_tz)
    
    if departure_time.tzinfo is None:
        departure_time = israel_tz.localize(departure_time)
    
    if departure_time <= reference_time:
        departure_time = reference_time + timedelta(minutes=5)

    unix_time = int(departure_time.timestamp())
    
    url = f"https://maps.googleapis.com/maps/api/directions/json?origin={origin}&destination={destination}&departure_time={unix_time}&key={api_key}"
    
    logging.info(f"API Request URL: {url}")
    
    response = requests.get(url)
    data = response.json()
    
    logging.info(f"API Response: {data}")
    
    if data['status'] == 'OK':
        travel_time = data['routes'][0]['legs'][0]['duration']['value']
        return travel_time
    else:
        error_message = f"Error fetching data from Google Maps API: {data['status']}"
        if 'error_message' in data:
            error_message += f" - {data['error_message']}"
        logging.error(error_message)
        raise Exception(error_message)


def get_active_drivers():
    active_crews = AmbulanceCrew.objects.filter(
        ambulance_id__status=1,
        end_time__isnull=True
    ).select_related('driver_id')
    
    return [get_object_or_404(BUser,pk = crew.driver_id) for crew in active_crews]


def get_active_rides_by_driver():
    # Get all active drivers
    drivers = get_active_drivers()
    
    # Initialize a dictionary to store rides by driver
    driver_rides = defaultdict(list)
    
    # Get all active rides and sort them by drop_off_time
    active_rides = CustomerRide.objects.filter(status=1).order_by('drop_of_time')
    
    # Iterate over the active rides and categorize them by driver
    for ride in active_rides:
        driver = ride.driver_id
        if driver in drivers:
            ride_info = {
                'origin': ride.pick_up_location,
                'destination': ride.drop_of_location,
                'start_time': ride.pick_up_time,
                'finish_time': ride.drop_of_time
            }
            driver_rides[driver].append(ride_info)

    return driver_rides


def can_fit_ride_between(start_time, end_time, new_ride_start, new_ride_end):
    return start_time > new_ride_end or end_time < new_ride_start


def can_driver_accommodate_ride(driver, new_ride, api_key):
    rides = get_active_rides_by_driver().get(driver, [])
    start_location = driver.current_location or "Kakal St 7, Petah Tikva"
    
    travel_time_to_pickup = get_travel_time(api_key, start_location, new_ride.pick_up_location, timezone.now())
    earliest_arrival = max(timezone.now() + timedelta(seconds=travel_time_to_pickup), new_ride.pick_up_time)
    
    possible_times = [earliest_arrival]
    handeling_time = timedelta(minutes=15)
    for i, ride in enumerate(rides):
        if i == 0:
            possible_times.append(earliest_arrival)
        else:
            prev_ride = rides[i-1]
            travel_time = get_travel_time(api_key, prev_ride['destination'], new_ride.pick_up_location, prev_ride['finish_time'] + handeling_time)
            possible_times.append(max(prev_ride['finish_time'] + timedelta(seconds=travel_time), new_ride.pick_up_time))
        
        travel_time2 = get_travel_time(api_key, new_ride.pick_up_location, new_ride.drop_of_location, possible_times[-1])
        new_ride_drop_of = possible_times[-1] + timedelta(seconds=travel_time2)
        
        next_ride_pick_up = ride['start_time']
        travel_time3 = get_travel_time(api_key, new_ride.drop_of_location, ride['origin'], new_ride_drop_of + handeling_time)
        next_ride_arrival = new_ride_drop_of + timedelta(seconds=travel_time3)
        
        if next_ride_arrival > next_ride_pick_up:
            possible_times[-1] = None
    
    if rides:
        last_ride = rides[-1]
        travel_time = get_travel_time(api_key, last_ride['destination'], new_ride.pick_up_location, last_ride['finish_time'] + handeling_time)
        possible_times.append(max(last_ride['finish_time'] + timedelta(seconds=travel_time), new_ride.pick_up_time))
    
    valid_times = [time for time in possible_times if time is not None]
    
    if valid_times:
        return True, min(valid_times)
    else:
        return False, None



def get_best_drivers_for_request(customer_request, api_key):
    preferred_driver = None
    if customer_request.have_preferred_driver:
        preferred_driver = get_object_or_404(BUser, pk=customer_request.preferred_driver)
    best_drivers = []

    active_drivers = get_active_drivers()
    
    def process_driver(driver):
        can_accommodate, arrival_time = can_driver_accommodate_ride(driver, customer_request, api_key)
        if can_accommodate:
            return driver, arrival_time
        return None
    
    with ThreadPoolExecutor() as executor:
        futures = [executor.submit(process_driver, driver) for driver in active_drivers]
        for future in as_completed(futures):
            result = future.result()
            if result:
                if preferred_driver and result[0] == preferred_driver:
                    best_drivers.insert(0, result)
                else:
                    best_drivers.append(result)
                if (preferred_driver and len(best_drivers) == 3) or (not preferred_driver and len(best_drivers) == 2):
                    break

    best_drivers.sort(key=lambda x: x[1])
    
    if preferred_driver:
        return best_drivers[:3] if len(best_drivers) >= 3 else best_drivers
    else:
        return best_drivers[:2]


def plan_a_ride(request):
    if request.method == 'POST':
        c_form = NewCustomerForm(request.POST)
        c_r_form = CustomerRequestForm(request.POST)
        
        if c_form.is_valid() and c_r_form.is_valid():
            new_customer = c_form.save()
            new_customer_request = c_r_form.save(commit=False)
            new_customer_request.customer_id = new_customer
            new_customer_request.save()
            print(f"Redirecting to pick_a_driver with request_id: {new_customer_request.id}")
            return redirect('pick_a_driver', request_id=new_customer_request.id)
        # else:
        #     print("Form is not valid")
        #     print(c_form.errors)
        #     print(c_r_form.errors)
    else:
        c_form = NewCustomerForm()
        c_r_form = CustomerRequestForm()
    
    return render(request, 'plan_a_ride.html', {'c_form': c_form, 'c_r_form': c_r_form})

def pick_a_driver(request, request_id):
    customer_request = get_object_or_404(CustomerRequest, pk=request_id)
    api_key = settings.GOOGLE_API_KEY
    best_drivers = get_best_drivers_for_request(customer_request, api_key)
    
    # Store arrival times in session
    request.session['driver_arrival_times'] = {str(driver.id): arrival_time.isoformat() for driver, arrival_time in best_drivers}
    time_now = timezone.now()
    context = {
        'customer_request': customer_request,
        'best_drivers': best_drivers,
        'time_now': time_now,
    }
    return render(request, 'pick_a_driver.html', context)


def complete_ride(request, request_id, driver_id):
    customer_request = get_object_or_404(CustomerRequest, pk=request_id)
    driver = get_object_or_404(BUser, pk=driver_id)
    
    crew = AmbulanceCrew.objects.filter(driver_id=driver, end_time__isnull=True).first()
    ambulance = crew.ambulance_id
    
    # Get the earliest arrival time from the session
    earliest_arrival = request.session.get('driver_arrival_times', {}).get(str(driver_id))
    if earliest_arrival:
        earliest_arrival = datetime.fromisoformat(earliest_arrival)
        if not timezone.is_aware(earliest_arrival):
            earliest_arrival = timezone.make_aware(earliest_arrival)
    # Clean the session variable
    if 'driver_arrival_times' in request.session:
        del request.session['driver_arrival_times']
    
    pick_up_time = earliest_arrival
    travel_time = get_travel_time(settings.GOOGLE_API_KEY, customer_request.pick_up_location, customer_request.drop_of_location, pick_up_time)
    drop_of_time = pick_up_time + timedelta(seconds=travel_time)
    
    CustomerRide.objects.create(
        customer_id=customer_request.customer_id,
        Am_id=ambulance,
        driver_id=driver,
        customer_req=customer_request,
        pick_up_location=customer_request.pick_up_location,
        drop_of_location=customer_request.drop_of_location,
        pick_up_time=pick_up_time,
        drop_of_time=drop_of_time,
        number_of_stuff_needed=customer_request.two_stuff_needed + 1,
        status=1,
    )
    
    return redirect('home')


def rides(request):
    ride_type = request.GET.get('type', 'current')
    filter_date = request.GET.get('date')

    form = DateFilterForm(initial={'date': filter_date or datetime.now().date()})

    if filter_date:
        filter_date = datetime.strptime(filter_date, '%Y-%m-%d').date()
    else:
        filter_date = datetime.now().date()

    if ride_type == 'current':
        rides = CustomerRide.objects.filter(status=3).order_by('pick_up_time')
    elif ride_type == 'scheduled':
        rides = CustomerRide.objects.filter(status=1, pick_up_time__date=filter_date).order_by('pick_up_time')
    else:  # completed
        rides = CustomerRide.objects.filter(status=2, pick_up_time__date=filter_date).order_by('pick_up_time')
    
    columns = ['Driver', 'Patient Name', 'Request Pick Up Time', 'Ride Pick Up Time']
    if ride_type == 'current':
        columns.append('Actions')
    elif ride_type == 'scheduled':
        columns[3] = 'Pick Up Location'
        columns.append('Actions')

    context = {
        'rides': rides,
        'columns': columns,
        'ride_type': ride_type,
        'filter_date': filter_date,
        'form': form,
    }
    return render(request, 'rides.html', context)



def generate_report(request):
    if request.method == 'POST':
        report_type = request.POST.get('report_type')
        from_date = request.POST.get('from_date')
        to_date = request.POST.get('to_date')
        export_format = request.POST.get('export_format')
        today = date.today().strftime("%d-%m-%Y")
        filename = f"ride report {today}.xlsx"

        # Query rides based on report type and date range
        rides = CustomerRide.objects.filter(pick_up_time__range=[from_date, to_date])
        if report_type == 'private':
            rides = rides.filter(customer_id__customer_type='private')
        elif report_type == 'business':
            rides = rides.filter(customer_id__customer_type='business')
        # Load the template
        template_path = os.path.join(settings.BASE_DIR, 'BubblanceApp/static/report_template.xlsx')
        workbook = load_workbook(template_path)
        sheet = workbook.active
        for index, ride in enumerate(rides, start=2):
            sheet.cell(row=index, column=1, value=ride.pick_up_time.date())
            sheet.cell(row=index, column=2, value=ride.customer_id.institution_id.in_institution.institution_name if ride.customer_id.institution_id and ride.customer_id.institution_id.in_institution else '')
            sheet.cell(row=index, column=3, value=ride.customer_id.institution_id.institution_name if ride.customer_id.institution_id else '')
            sheet.cell(row=index, column=4, value=ride.customer_id.patient_name)
            sheet.cell(row=index, column=5, value=ride.pick_up_location)
            sheet.cell(row=index, column=6, value=ride.drop_of_location)
        temp_excel_path = os.path.join(settings.TEMP_DIR, f"temp_report_{today}.xlsx")
        workbook.save(temp_excel_path)

        if export_format == 'excel':
            with open(temp_excel_path, 'rb') as f:
                response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
        elif export_format == 'pdf':
            pythoncom.CoInitialize()
            # Convert Excel to PDF
            excel = win32com.client.Dispatch("Excel.Application")
            wb = excel.Workbooks.Open(temp_excel_path)
            temp_pdf_path = os.path.join(settings.TEMP_DIR, f"temp_report_{today}.pdf")
            wb.SaveAs(temp_pdf_path, FileFormat=57)  # 57 is the code for PDF
            wb.Close()
            excel.Quit()
            pythoncom.CoUninitialize()
   
            with open(temp_pdf_path, 'rb') as f:
                response = HttpResponse(f.read(), content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="{filename.replace(".xlsx", ".pdf")}"'

        # Clean up temporary files
        os.remove(temp_excel_path)
        if export_format == 'pdf':
            os.remove(temp_pdf_path)

        return response

    return render(request, 'generate_report.html')


def driver_home(request):
    if request.session.pop('navigation_initiated', False):
        # Clear the session variable and redirect to driver home
        return redirect(reverse('driver_home'))
    driver = request.user.buser
    page_type = request.GET.get('type', 'todo')
    selected_date = request.GET.get('date')
    active_crew = AmbulanceCrew.objects.filter(driver_id=driver, end_time__isnull=True).first()
    active_ambulance_ids = AmbulanceCrew.objects.filter(end_time__isnull=True).values_list('ambulance_id', flat=True)
    available_ambulances = Ambulance.objects.filter(status=1).exclude(ambulance_id__in=active_ambulance_ids)

    if selected_date:
        selected_date = datetime.strptime(selected_date, '%Y-%m-%d').date()
    else:
        selected_date = datetime.now().date()

    ongoing_ride = CustomerRide.objects.filter(driver_id=driver, status=3).first()

    if page_type == 'todo':
        active_rides = CustomerRide.objects.filter(driver_id=driver, status=1, pick_up_time__date=selected_date)
    else:
        finished_rides = CustomerRide.objects.filter(driver_id=driver, status=2, pick_up_time__date=selected_date)

    context = {
        'active_crew': active_crew,
        'available_ambulances': available_ambulances,
        'ongoing_ride': ongoing_ride,
        'rides': active_rides if page_type == 'todo' else finished_rides,
        'page_type': page_type,
        'selected_date': selected_date
    }
    return render(request, 'driver_home.html', context)


def ride_details(request, ride_id):
    ride = get_object_or_404(CustomerRide, cust_ride_id=ride_id)
    is_manager = False
    if request.user.buser.usertype == 2:
        is_manager = True 

    customer_link = None
    if is_manager:
        customer_link = request.build_absolute_uri(reverse('customer_ride_page'))
        customer_link += f'?token={ride.token}'  # You'll need to implement this method

    context = {
        'ride': ride,
        'is_manager': is_manager,
        'customer_link': customer_link,
    }
    return render(request, 'ride_details.html', context)


def start_ride(request, ride_id):
    ride = get_object_or_404(CustomerRide, cust_ride_id=ride_id)
    ride.status = 3  # Set status to ongoing
    ride.save()
    messages.success(request, "Ride started successfully.")
    return redirect('ride_details', ride_id=ride_id)


def finish_ride(request, ride_id):
    ride = get_object_or_404(CustomerRide, cust_ride_id=ride_id)
    ride.status = 2  # Set status to deactive
    ride.save()
    if request.user.buser.usertype == 2:  # Manager
        return redirect(reverse('rides'))
    else:  # Driver
        return redirect(reverse('driver_home'))

def customer_ride_page(request):
    token = request.GET.get('token')
    ride = get_object_or_404(CustomerRide, token=token)
    
    context = {
        'ride': ride,
    }
    return render(request, 'customer_ride.html', context)


def get_institution_address(request, institution_id):
    inst = get_object_or_404(Institution, inst_id=institution_id)
    try:
        return JsonResponse({'address': inst.institution_adress})
    except Institution.DoesNotExist:
        return JsonResponse({'error': 'Institution not found'}, status=404)
    
    
def start_shift(request):
    if request.method == 'POST':
        ambulance_id = request.POST.get('ambulance_id')
        driver = request.user.buser
        AmbulanceCrew.objects.create(
            ambulance_id_id=ambulance_id,
            driver_id=driver,
            start_time=timezone.now()
        )
        return redirect('driver_home')


def finish_shift(request):
    driver = request.user.buser
    active_crew = AmbulanceCrew.objects.filter(driver_id=driver, end_time__isnull=True).first()
    if active_crew:
        active_crew.end_time = timezone.now()
        active_crew.save()
    return redirect('driver_home')